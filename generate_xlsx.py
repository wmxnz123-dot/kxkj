import subprocess
import sys
import os

def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    install('openpyxl')
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "功能清单"

headers = ["模块", "功能名称", "功能描述", "使用端", "前置条件"]
data = [
    ["首页", "数据概览", "展示我的数据、自有数据、授权记录、我的服务数量", "App端", "登录并实名认证"],
    ["首页", "快捷服务", "提供我的数据、数据认领、自有数据、个人授权快速入口", "App端", "无"],
    ["首页", "应用服务", "展示教育、就业、健康、资产等分类服务入口", "App端", "无"],
    ["首页", "消息中心", "展示未读/已读系统消息、授权通知等", "App端", "无"],
    ["我的", "个人信息", "展示姓名、身份证号、住址，支持脱敏显示开关", "App端", "无"],
    ["我的", "个人画像", "展示用户头像及个人标签", "App端", "无"],
    ["我的", "数据仪表盘", "按社保、电子证照、医疗健康等分类展示数据资产", "App端", "无"],
    ["我的", "设置-使用说明", "查看平台的使用指南和帮助文档", "App端", "无"],
    ["我的", "设置-纠错历史", "查看用户发起的所有数据纠错记录及处理状态", "App端", "无"],
    ["我的", "设置-隐私政策", "查看平台的隐私保护政策及用户协议", "App端", "无"],
    ["我的", "设置-关于我们", "展示平台版本信息、联系方式等", "App端", "无"],
    ["我的", "设置-退出登录", "退出当前账号，清除本地登录状态", "App端", "无"],
    ["我的数据", "已认领数据目录", "列表展示已认领的数据目录，点击需人脸认证后查看详情", "App端", "已认领相关数据"],
    ["我的数据", "未认领数据目录", "列表展示系统推荐可认领的数据，支持一键认领", "App端", "无"],
    ["我的数据", "数据认领", "对推荐或搜索到的数据进行认领操作", "App端", "无"],
    ["数据详情", "数据详情查看", "展示具体的数据项及数值", "App端", "通过人脸认证"],
    ["数据详情", "数据纠错", "对有误的数据发起纠错申请，填写纠错原因和证明材料", "App端", "无"],
    ["数据详情", "数据更新", "触发数据源同步最新数据", "App端", "无"],
    ["自有数据", "自有数据列表", "展示用户自行上传或填写的非政务数据", "App端", "无"],
    ["自有数据", "自有数据上传", "支持用户拍照或本地上传证明材料", "App端", "无"],
    ["自有数据", "自有数据查看", "查看已上传的自有数据详情及附件", "App端", "无"],
    ["自有数据", "自有数据编辑", "修改已上传的自有数据信息", "App端", "无"],
    ["自有数据", "自有数据删除", "删除不再需要的自有数据", "App端", "无"],
    ["个人授权", "待授权列表", "展示第三方应用或机构发起的授权申请", "App端", "无"],
    ["个人授权", "授权详情查看", "查看第三方应用请求授权的具体数据项和用途", "App端", "无"],
    ["个人授权", "同意授权", "同意第三方的授权申请", "App端", "无"],
    ["个人授权", "拒绝授权", "拒绝第三方的授权申请", "App端", "无"],
    ["个人授权", "已授权列表", "展示当前生效中的授权，支持取消授权", "App端", "无"],
    ["个人授权", "取消授权", "提前终止或撤销正在生效的授权", "App端", "无"],
    ["个人授权", "授权记录", "查看所有的历史授权和取消授权记录", "App端", "无"],
    ["数据服务", "定制化", "按需为用户提供定制化的数据服务场景", "App端", "无"],
    ["AI赋能", "智能助理", "大模型对话", "以用户的个人数据空间为私有知识库进行对话交互", "App端", "无"],
    ["AI赋能", "智能助理", "简历/报告生成", "根据指令自动提取个人数据并生成排版好的PDF简历或报告", "App端", "无"],
    ["AI赋能", "智能解读", "数据详情解读", "通过大模型对体检报告、社保等复杂数据进行白话解读和健康建议", "App端", "无"],
    ["AI赋能", "智能审核", "纠错材料预审", "自动识别用户上传的证明材料，预判是否与纠错诉求相符", "App/后台", "无"],
    ["AI赋能", "智能录入", "自有数据提取", "用户上传PDF或照片后，AI自动提取关键字段并结构化存入数据库", "App端", "无"],
    ["AI赋能", "智能风控", "异常授权拦截", "检测非工作时间或高频次的异常拉取行为，自动切断授权并发送通知", "系统后台", "无"],
    ["AI赋能", "智能客服", "操作引导答疑", "将使用说明投喂给大模型，为用户提供精准的操作步骤引导", "App端", "无"],
    ["后台管理", "用户管理", "查看App端注册用户信息及实名状态", "Admin端", "管理员权限"],
    ["后台管理", "数据目录管理", "配置可供认领和接入的数据源目录", "Admin端", "管理员权限"],
    ["后台管理", "纠错审核", "审核用户提交的数据纠错申请，推进处理进度", "Admin端", "管理员权限"],
    ["后台管理", "授权模板管理", "配置不同场景下的数据授权范围和模板", "Admin端", "管理员权限"],
    ["后台管理", "系统设置", "平台基础配置、角色权限分配等", "Admin端", "管理员权限"]
]

# Write headers
ws.append(headers)

# Style headers
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Write data
for row in data:
    ws.append(row)

# Adjust column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 20

# Align data
for row in ws.iter_rows(min_row=2, max_col=5, max_row=len(data)+1):
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# Save the file
file_name = "个人专属数据空间_功能清单.xlsx"
wb.save(file_name)
print(f"Success: {file_name} generated.")
